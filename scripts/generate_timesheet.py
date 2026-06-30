"""
Generate monthly timesheet Excel from the May/June template.
Tasks are auto-filled from git commit history (two lines per day).

Usage:
  py scripts/generate_timesheet.py --month 7 --year 2026
  py scripts/generate_timesheet.py --month 7          # defaults to current year
"""
from __future__ import annotations

import argparse
import calendar
import re
import subprocess
import sys
from collections import defaultdict
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl

WEEKDAYS = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]

TEMPLATE = Path(r"c:\Users\hp\OneDrive\Desktop\Time sheets\Time Sheet_May .xlsx")
OUTPUT_DIR = Path(r"c:\Users\hp\OneDrive\Desktop\Time sheets")
REPO_ROOT = Path(__file__).resolve().parents[1]
HOURS_PER_DAY = 4

MONTH_NAMES = [
    "",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]


def git_commits_by_date(year: int, month: int) -> dict[date, list[str]]:
    """Return commit messages grouped by commit date for the given month."""
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)

    result = subprocess.run(
        [
            "git",
            "log",
            f"--since={start.isoformat()}",
            f"--until={end.isoformat()}",
            "--format=%ad|%s",
            "--date=short",
        ],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        print(result.stderr, file=sys.stderr)
        return {}

    by_date: dict[date, list[str]] = defaultdict(list)
    for line in result.stdout.splitlines():
        if "|" not in line:
            continue
        day_str, message = line.split("|", 1)
        try:
            y, m, d = (int(x) for x in day_str.split("-"))
            by_date[date(y, m, d)].append(message.strip())
        except ValueError:
            continue
    return by_date


def clean_message(message: str) -> str:
    """Turn a commit subject into a readable task line."""
    text = message.strip()
    text = re.sub(r"^code upgrade on\s+", "", text, flags=re.I)
    text = re.sub(r"^code upgrade to\s+", "", text, flags=re.I)
    text = re.sub(r"^code upgrade in\s+", "", text, flags=re.I)
    text = re.sub(r"^code upgrade\s*", "", text, flags=re.I)
    text = re.sub(r"^fix:\s*", "Fix ", text, flags=re.I)
    text = re.sub(r"^fix\s+", "Fix ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text)
    if text:
        text = text[0].upper() + text[1:]
    if text and text[-1] not in ".!?":
        text += "."
    return text


def summarize_commits(commits: list[str]) -> str:
    """Build two task lines from one day's commits."""
    unique: list[str] = []
    seen: set[str] = set()
    for msg in commits:
        key = msg.lower().strip()
        if key in seen:
            continue
        seen.add(key)
        unique.append(clean_message(msg))

    if not unique:
        return ""

    if len(unique) == 1:
        return f"{unique[0]}\nTesting, review, and follow-up fixes."

    line1 = unique[0]
    line2 = unique[1]
    if len(unique) > 2:
        extra = unique[2]
        if len(extra) < 80:
            line2 = f"{line2.rstrip('.')}; {extra[0].lower() + extra[1:] if len(extra) > 1 else extra}"

    return f"{line1}\n{line2}"


def copy_header(may_ws, target_ws) -> None:
    for row in may_ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=4):
        for cell in row:
            target = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target.font = copy(cell.font)
                target.fill = copy(cell.fill)
                target.border = copy(cell.border)
                target.alignment = copy(cell.alignment)
                target.number_format = cell.number_format

    for col in ("A", "B", "C", "D"):
        target_ws.column_dimensions[col].width = may_ws.column_dimensions[col].width


def generate_timesheet(year: int, month: int) -> Path:
    month_name = MONTH_NAMES[month]
    output_path = OUTPUT_DIR / f"Time Sheet_{month_name} .xlsx"

    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")

    commits_by_date = git_commits_by_date(year, month)
    days_in_month = calendar.monthrange(year, month)[1]

    wb = openpyxl.load_workbook(TEMPLATE)
    may_ws = wb["May"] if "May" in wb.sheetnames else wb.active

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = month_name
    copy_header(may_ws, ws)

    date_fmt = may_ws["A7"].number_format if may_ws["A7"].value else "dd/mm/yyyy"

    for day in range(1, days_in_month + 1):
        row = 6 + day
        current = date(year, month, day)
        weekday = WEEKDAYS[current.weekday()]
        is_sunday = current.weekday() == 6

        ws.cell(row=row, column=1, value=current)
        ws.cell(row=row, column=1).number_format = date_fmt
        ws.cell(row=row, column=2, value=weekday)

        day_commits = commits_by_date.get(current, [])
        if day_commits:
            ws.cell(row=row, column=3, value=summarize_commits(day_commits))
            ws.cell(row=row, column=4, value=HOURS_PER_DAY)
        elif is_sunday:
            ws.cell(row=row, column=3, value="Weekend")
            ws.cell(row=row, column=4, value=None)
        else:
            ws.cell(row=row, column=3, value=None)
            ws.cell(row=row, column=4, value=None)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)

    filled = sum(1 for d in range(1, days_in_month + 1) if commits_by_date.get(date(year, month, d)))
    print(f"Saved: {output_path}")
    print(f"Days with auto-filled tasks: {filled}/{days_in_month}")
    if filled < days_in_month:
        print("Re-run this script anytime during the month to refresh tasks from new git commits.")
    return output_path


def main() -> None:
    today = date.today()
    parser = argparse.ArgumentParser(description="Generate monthly KITERP timesheet from git history.")
    parser.add_argument("--month", type=int, required=True, help="Month number (1-12)")
    parser.add_argument("--year", type=int, default=today.year, help="Year (default: current year)")
    args = parser.parse_args()

    if not 1 <= args.month <= 12:
        parser.error("--month must be between 1 and 12")

    generate_timesheet(args.year, args.month)


if __name__ == "__main__":
    main()
